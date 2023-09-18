#--------------------------------------------------------------------------------------
#
#  Copyright: (c) 2020 Bentley Systems, Incorporated. All rights reserved. 
#
#--------------------------------------------------------------------------------------


from ram_concept.cad_manager import CadManager
from ram_concept.elements import ColumnElement
from ram_concept.elements import WallElementGroup
from ram_concept.element_layer import ElementLayer
from ram_concept.force_loading_layer import ForceLoadingLayer
from ram_concept.line_segment_2D import LineSegment2D
from ram_concept.loading_layer import LoadingCause
from ram_concept.loading_layer import LoadingType
from ram_concept.model import Model
from ram_concept.point_2D import Point2D
from ram_concept.point_3D import Point3D
from ram_concept.result_layers import ReactionContext

import time
import openpyxl
import sys

def lb_to_kip(value):
    return value / 1000

def in_to_ft(value):
    return value / 12



def get_reactions(model: Model):
      

    ######################################################################      VARIABLES     ##################################################################

    run_from_VBA = False

    if run_from_VBA:
        ### Get variables from VBA
        start_level, end_level, loading_type, max_height, model_filepath, name_DL, name_LL, name_trib = [sys.argv[i] for i in range(2,10)]
    else:
        ### for debugging without VBA
        start_level, end_level, loading_type, max_height, model_filepath, name_DL, name_LL, name_trib = 4, 4, "RESIDENTIAL", 11, "H:\Internal Innovation\RAM to Takedown\RAM\Dining Hall 2.cpt", "All Dead LC", "Live+Soil (total Live Load reactions)", "Trib"


    # Cast number types
    start_level = int(start_level)
    end_level = int(end_level)
    max_height = float(max_height)

    # Create results list
    takedown_list = []

    ######################################################################      COMBO DATA     ##################################################################

    cad_manager = model.cad_manager
    element_layer = cad_manager.element_layer
    
    # get all the loadings
    loadings = cad_manager.force_loading_layers

    # get all the load combos
    load_combos = cad_manager.load_combo_layers

    # combine the load combo layers and the loading layers, as we want reactions for both
    loadings_and_combos = loadings + load_combos

    #set load combo names from userform
    # name_DL = "All Dead LC"
    # name_LL = "Live+Soil (total Live Load reactions)"
    # name_trib = "Trib"

    #name_DL = cmb_dead
    #name_LL = cmb_live
    #name_trib = cmb_trib

    # get relevant load combo objects from list of loading layers
    
    combo_dead = "Dead load combo not found in model."
    combo_live = "Live load combo not found in model."
    combo_trib = "Trib load case not found in model."
    
    combo_list = [combo_dead, combo_live, combo_trib]

    # Match combo names with combo layer in model
    for combo in load_combos:
        combo_name = combo.name
        if combo_name == name_DL:
            combo_dead = combo
        elif combo_name == name_LL:
            combo_live = combo
    
    for combo in loadings:
        if combo.name == name_trib or "trib" in combo.name.lower():
            combo_trib = combo

    # Check if any combo was not found
    if type(combo_dead) == str:
        return combo_dead
    elif type(combo_live) == str:
        return combo_live
    elif type(combo_trib) == str:
        return combo_trib
    

    ######################################################################      COLUMN DATA     ##################################################################
    
    # loop through column elements, add column data to list
    for column_element in element_layer.column_elements_below:
        # reaction = loading_or_combo.column_reaction(column_element, ReactionContext.STANDARD)
        # location = column_element.location
        # print("{0:7.2f} {1:7.2f} {2:9.2g} {3:9.2g} {4:9.2g} {5:9.2g} {6:9.2g} {7:9.2g} {8:9.2g}".format(location.x, location.y, reaction.x, reaction.y, reaction.z, reaction.rot_x, reaction.rot_y, column_element.b, column_element.d))

        t1 = time.time()

        location = column_element.location

        # get column reactions from 3 combos
        rxn_DL = combo_dead.column_reaction(column_element, ReactionContext.STANDARD).z
        rxn_LL = combo_live.column_reaction(column_element, ReactionContext.STANDARD).z
        rxn_trib = combo_trib.column_reaction(column_element, ReactionContext.STANDARD).z

        t2 = time.time()
        #print("Combo time: " + str(t2-t1))

        # Compile list of current column's data, convert to desired units
        col_data = [in_to_ft(location.x), in_to_ft(location.y), lb_to_kip(rxn_DL), lb_to_kip(rxn_LL), in_to_ft(column_element.height), column_element.b, column_element.d, lb_to_kip(rxn_trib)]

        t3 = time.time()
        #print("Col_data time: " + str(t3-t2))

        # add current column data list to full takedown list
        takedown_list.append(col_data)
     #   print(len(takedown_list))

        t4 = time.time()
        #print("Takedown time: " + str(t4-t3))

      #  print()
      #  print()


    ######################################################################      WALL DATA     ##################################################################

    


    #wall_group_numbers = ["#" + str(i) for i in range(1,len(element_layer.wall_element_groups_below)+1)]
    
    """     ### Method 1

    wall_thickness_dict = {}

    t5 = time.time()

    i = 1
    for wall_element in element_layer.wall_elements_below:
        #print(wall_element.name == "(#1)")
        elem_name = "(#" + str(i) + ")"
        if wall_element.name == elem_name:
            wall_thickness_dict[elem_name] = [wall_element.thickness, wall_element.height/12]
            i += 1
    
    t6 = time.time()
    print("Wall search time: " + str(t6-t5)) """


    ### Method 2

    wall_thickness_dict = {}
    t6 = time.time()

    for wall_element in element_layer.wall_elements_below:
        if not wall_element.name in wall_thickness_dict:
            wall_thickness_dict[wall_element.name] = [wall_element.thickness, wall_element.height/12]

    t7 = time.time()
    #print("Build wall dict 2: " + str(t7-t6))
    #username = input("Continue?")
    
    
    """ ### Method 3
    wall_thickness_dict = {}
    t7 = time.time()

    wall_group_numbers = ["(#" + str(i) + ")" for i in range(1,len(element_layer.wall_element_groups_below)+1)]
    
    for wall_element in element_layer.wall_elements_below:                                      # Loop through wall elements
        elem_name = wall_element.name
        if elem_name in wall_group_numbers:                                                     # If element name is still in wall_group_numbers:
            wall_height = wall_element.height/12
            if wall_height <= hLimit:                                                               # If element height is within limit
                wall_thickness_dict[elem_name] = [wall_element.thickness, wall_height]                  # Add to dictionary
            wall_group_numbers.remove(elem_name)                                                    # Remove element name from wall_group_numbers

    t8 = time.time()
    print("Build wall dict 3: " + str(t8-t7)) """
 


    #wall_element_dict = {}
    #for wall_element in element_layer.wall_elements_below:
    #    wall_element_dict[wall_element.name] = [wall_element.thickness, wall_element.height/12]
    #
    #t8 = time.time()
    #print("Wall dict time: " + str(t8-t7))



    #print(element_layer.wall_elements_below)
    #
    #username = input("Continue?")
    #print("answer: " + username) 


    t9 = time.time()

    #### loop through wall element groups, add wall data to list
    for wall_element_group in element_layer.wall_element_groups_below:
       
       # OLD WALL THICKNESS SEARCH
        # loop through wall elements to find one in this wall element group and get thickness and height
        #for wall_element in element_layer.wall_elements_below:
        #    if wall_element.name == wall_element_group.name:
        #        b = wall_element.thickness
        #        h = wall_element.height
        #      #  print("Wall element name: " + str(wall_element.name))
        #      #  print("Thickness: " + str(b) + " , Height: " + str(h))
        #        break


        # Check is wall met height limit and therefore is in wall thickness dict
        #if wall_element_group.name in wall_thickness_dict:

        # NEW WALL THICKNESS SET
        group_name = wall_element_group.name
        b = wall_thickness_dict[group_name][0]
        h = wall_thickness_dict[group_name][1]


        if h <= max_height:
        
            # reaction = loading_or_combo.wall_group_reaction(wall_element_group, ReactionContext.STANDARD)
            #name = wall_element_group.name
            #centroid = wall_element_group.centroid
            #angle = wall_element_group.reaction_angle
            #length = wall_element_group.total_length
            #area = wall_element_group.total_area

            rxn_DL = combo_dead.wall_group_reaction(wall_element_group, ReactionContext.STANDARD).z
            rxn_LL = combo_live.wall_group_reaction(wall_element_group, ReactionContext.STANDARD).z
            rxn_trib = combo_trib.wall_group_reaction(wall_element_group, ReactionContext.STANDARD).z

            centroid = wall_element_group.centroid
            d = wall_element_group.total_length
         #   print("d: " + str(d))
            
            
            # compile wall data into list
            col_data = [in_to_ft(centroid.x), in_to_ft(centroid.y), lb_to_kip(rxn_DL), lb_to_kip(rxn_LL), h, b, d, lb_to_kip(rxn_trib)]
            #col_data = [ '%.3f' % elem for elem in col_data]

            # add current wall data list to full takedown list
            takedown_list.append(col_data)


    t10 = time.time()
    #print("Wall calcs: " + str(t10-t9))

    #print(len(takedown_list))



    # RETURN 
    return takedown_list





##Write text file with level data
    #takedown_headers = ['x','y','DL','LL','H','b','d','Trib']
    #outfile = 'takedownList.txt'
    #with open(outfile, 'w') as file:
    #    file.writelines('\t'.join(takedown_headers) + '\n')
    #    file.writelines('\t'.join(str(round(cell,2)) for cell in element) + '\n' for element in takedown_list)



    
# Openpyxl
""" # # Get userform data from Excel
    wb = openpyxl.load_workbook(wbPath)
    ws = wb['AutoTD']

    # data_list = [ws.cell(row=i,column=2).value for i in range(14,22)]
    # start_level, end_level, loading_type, max_height, model_path, cmb_dead, cmb_live, cmb_trib = [data_list[i] for i in range(0,len(data_list))]
    # print("start: " + str(start_level))
    # print("end: "+ str(end_level))
    # print("loading type: "+ str(loading_type))
    # print(data_list)    

    ########### WRITE TAKEDOWN_LIST TO EXCEL FILE
    # Get position to paste data
    num_levels = 0
    for i in range(5,106):
        cell_value = ws.cell(row=i, column=1)

        if cell_value == "" or cell_value == start_level:
            break

        num_levels =+ 1
    
    # Get range to paste takedown_list
    start_column = int(3 + 8 * num_levels)
    start_row = int(4)
    end_column = start_column + 7
    end_row = start_row + len(takedown_list) - 1

    # Paste import level headers
    ws.cell(row=5 + num_levels, column=1).value = start_level
    ws.cell(row=1, column=start_column).value = start_level
    ws.cell(row=1, column=start_column+1).value = loading_type
 """











#       # for combo in loadings_and_combos:
        #     if combo.name in combo_list:
        #         reaction = combo.column_reaction(column_element, ReactionContext.STANDARD)
        #         col_data.append(reaction.z)
        
        # col_data.insert(4, column_element.height)
        # col_data.insert(5, column_element.b)
        # col_data.insert(6, column_element.d)

"""     wanted_combos = []
    #FIND DESIRED LOAD COMBOS
    for loading_or_combo in loadings_and_combos:
        if loading_or_combo.name in combo_list:
            wanted_combos.append(loading_or_combo)
            print(loading_or_combo.name + " combo acquired") """


        # this works but is slow
        # for combo in loadings_and_combos:
        #     if combo.name == combo_DL:
        #         rxn_DL = combo.column_reaction(column_element, ReactionContext.STANDARD).z
        #     elif combo.name == combo_LL:
        #         rxn_LL = combo.column_reaction(column_element, ReactionContext.STANDARD).z
        #     elif combo.name == combo_trib:
        #         rxn_trib = combo.column_reaction(column_element, ReactionContext.STANDARD).z



    # # remove the hyperstatic loading
    # for loading in loadings:
    #     if loading.loading_type.cause == LoadingCause.HYPERSTATIC:
    #         loadings.remove(loading)
    #         break




#
"""         # # loop through all the loadings and combos and get their reactions
            # for loading_or_combo in loadings_and_combos:
                # put a pleasant header
                header = loading_or_combo.name + " REACTIONS"
                print(header)
                print("*" * len(header))
                print()

                # handle column reactions
                print("Column Reactions")
                print("----------------")
                print("    x       y       Fx        Fy        Fz        Mx        My        b        d")
                print("------------------------------------------------------------------")
    #print("Units: " + str(model.units.get_units))
                for column_element in element_layer.column_elements_below:
                    reaction = loading_or_combo.column_reaction(column_element, ReactionContext.STANDARD)
                    location = column_element.location
                    print("{0:7.2f} {1:7.2f} {2:9.2g} {3:9.2g} {4:9.2g} {5:9.2g} {6:9.2g} {7:9.2g} {8:9.2g}".format(location.x, location.y, reaction.x, reaction.y, reaction.z, reaction.rot_x, reaction.rot_y, column_element.b, column_element.d))

                    #TAKEDOWN ARRAY
                    takedown_list.append([location.x, location.y ])
                    #CHECK IF COLUMN EXISTS IN DICTIONARY




                # add a blank line between columns and walls
                print()

                # handle wall reactions
                print("Wall Reactions")
                print("--------------")
                print(" name      x       y       z    angle   length   area      Fx        Fy        Fz        Mx        My        Mz")
                print("----------------------------------------------------------------------------------------------------------------")
                for wall_element_group in element_layer.wall_element_groups_below:
                    reaction = loading_or_combo.wall_group_reaction(wall_element_group, ReactionContext.STANDARD)
                    name = wall_element_group.name
                    centroid = wall_element_group.centroid
                    angle = wall_element_group.reaction_angle
                    length = wall_element_group.total_length
                    area = wall_element_group.total_area
                    print("{0:5} {1:7.2f} {2:7.2f} {3:7.2f} {4:7.2f} {5:7.2f} {6:7.2f} {7:9.2g} {8:9.2g} {9:9.2g} {10:9.2g} {11:9.2g} {12:9.2g}".format(
                        name, centroid.x, centroid.y, centroid.z, angle, length, area, reaction.x, reaction.y, reaction.z, reaction.rot_x, reaction.rot_y, reaction.rot_z)) """